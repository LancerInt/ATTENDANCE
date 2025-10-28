import streamlit as st
import pandas as pd
import numpy as np
import re
import tempfile
from datetime import datetime, time
from io import BytesIO
from openpyxl.utils import get_column_letter

# ======================================================
# Utility Functions
# ======================================================
def clean_id(x):
    s = str(x).strip().upper()
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s

def to_time(v):
    """Extract time from text like '23:04:00 - O' or '07:16:00'"""
    if pd.isna(v):
        return None
    s = str(v)
    s = re.sub(r"[^0-9:]", "", s).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None

def find_emp_col(df):
    possible = ["pay code", "emp code", "employee code", "empid", "emp id", "code"]
    safe_cols = [str(c).strip() for c in df.columns]
    for name in possible:
        for i, c in enumerate(safe_cols):
            if name in c.lower():
                return df.columns[i]
    return df.columns[0]

def dedupe_columns_inplace(df):
    df.columns = [str(c) for c in df.columns]
    mask = ~pd.Index(df.columns).duplicated(keep="first")
    return df.loc[:, mask]

# ======================================================
# Comparison Logic
# ======================================================
def compare_files(attendance_bytes, bio1_bytes, bio2_bytes):
    # Save temporary copies
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as att_temp:
        att_temp.write(attendance_bytes); att_temp.flush(); att_path = att_temp.name
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as b1:
        b1.write(bio1_bytes); b1.flush(); bio1_path = b1.name
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as b2:
        b2.write(bio2_bytes); b2.flush(); bio2_path = b2.name

    # Attendance file
    xls = pd.ExcelFile(att_path)
    sheet_name = next((s for s in xls.sheet_names if "data" in s.lower() and "entry" in s.lower()), xls.sheet_names[0])
    manual_df = pd.read_excel(att_path, sheet_name=sheet_name)
    manual_df.columns = manual_df.columns.str.strip()

    date_col = next((c for c in manual_df.columns if "date" in c.lower()), None)
    if date_col:
        manual_df[date_col] = pd.to_datetime(manual_df[date_col], errors="coerce")

    # Biometric files
    bio1 = pd.read_excel(bio1_path, skiprows=1)
    bio2 = pd.read_excel(bio2_path, skiprows=1)
    bio1 = dedupe_columns_inplace(bio1)
    bio2 = dedupe_columns_inplace(bio2)
    bio1.columns = bio1.columns.str.strip()
    bio2.columns = bio2.columns.str.strip()

    # Detect EMP ID columns
    m_emp = next((c for c in manual_df.columns if "emp id" in c.lower()), "EMP ID")
    b_emp1, b_emp2 = find_emp_col(bio1), find_emp_col(bio2)

    manual_df[m_emp] = manual_df[m_emp].astype(str).map(clean_id)
    bio1[b_emp1] = bio1[b_emp1].astype(str).map(clean_id)
    bio2[b_emp2] = bio2[b_emp2].astype(str).map(clean_id)

    bio1 = bio1.replace(r"^\s*$|^-$", np.nan, regex=True)
    bio2 = bio2.replace(r"^\s*$|^-$", np.nan, regex=True)

    # Helper: collect punch times
    def get_punch_times(df, emp, id_col):
        sub = df[df[id_col] == emp]
        if sub.empty:
            return []
        r = sub.iloc[0]
        punches = []
        for c in df.columns:
            if "PUNCH" in str(c).upper() and pd.notna(r[c]):
                t = to_time(r[c])
                if t:
                    punches.append(t)
        punches.sort()
        return punches

    # Define shift windows for "Day/Morning"
    in_start = time(6, 45)
    in_end   = time(7, 30)
    out_start = time(15, 15)
    out_end   = time(15, 20)

    statuses = []
    shift_col = next((c for c in manual_df.columns if "shift" in c.lower()), "SHIFT")
    all_bio1_ids = set(bio1[b_emp1])

    for _, row in manual_df.iterrows():
        emp = row[m_emp]
        shift = str(row.get(shift_col, "")).strip().lower()
        if emp not in all_bio1_ids:
            statuses.append("No Punch")
            continue

        punches_day1 = get_punch_times(bio1, emp, b_emp1)
        punches_day2 = get_punch_times(bio2, emp, b_emp2)
        status = "No Punch"

        # ---------- FULL NIGHT ----------
        if "full" in shift or "fn" in shift:
            p1_day1 = punches_day1[0] if punches_day1 else None
            p1_day2 = punches_day2[0] if punches_day2 else None
            if p1_day1 and p1_day2:
                status = "Match"
            elif p1_day1 and not p1_day2:
                status = "Match"
            elif not p1_day1 and p1_day2:
                status = "Single Out Punch"
            else:
                status = "No Punch"

        # ---------- DAY / MORNING SHIFT ----------
        elif "day" in shift or "morning" in shift:
            if len(punches_day1) < 2:
                status = "No Punch"
            else:
                in_time, out_time = punches_day1[0].time(), punches_day1[-1].time()
                # Check IN time window
                if not (in_start <= in_time <= in_end):
                    status = "No Punch"
                else:
                    # Check OUT time
                    if out_time < out_start:
                        status = "Early"
                    elif out_start <= out_time <= out_end:
                        status = "Match"
                    else:
                        status = "No Punch"

        # ---------- OTHER SHIFTS (HF etc.) ----------
        else:
            n = len(punches_day1)
            def diff_mins(a, b): return abs((b - a).total_seconds()) / 60 if a and b else 9999
            if n == 0:
                status = "No Punch"
            elif n >= 4:
                t1, t4 = punches_day1[0], punches_day1[3]
                status = "Single In Punch" if diff_mins(t1, t4) <= 10 else "Match"
            elif n == 3:
                t1, t3 = punches_day1[0], punches_day1[2]
                status = "Single In Punch" if diff_mins(t1, t3) <= 10 else "Match"
            elif n == 2:
                t1, t2 = punches_day1[0], punches_day1[1]
                status = "Single In Punch" if diff_mins(t1, t2) <= 10 else "Match"
            elif n == 1:
                status = "Single In Punch"
            else:
                status = "Mismatch"

        statuses.append(status)

    manual_df["Status"] = statuses

    # Output Excel with date formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        manual_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        if date_col:
            col_idx = manual_df.columns.get_loc(date_col) + 1
            col_letter = get_column_letter(col_idx)
            for row in range(2, len(manual_df) + 2):
                ws[f"{col_letter}{row}"].number_format = "yyyy-mm-dd"
            ws.column_dimensions[col_letter].width = 15
    output.seek(0)
    return output, manual_df

# ======================================================
# Streamlit UI
# ======================================================
st.set_page_config(page_title="Attendance Comparator", page_icon="🌞", layout="centered")
st.title("🌞 Attendance Comparator (Fullnight + Morning Shift Logic)")
st.markdown("""
Upload three files:  
1️⃣ **Attendance DailyEntry File (Data_Entry)**  
2️⃣ **Biometric File – Day 1 (Night Punches)**  
3️⃣ **Biometric File – Day 2 (Morning Punches)**  

Morning shift timing logic added:  
- IN valid: 06:45 → 07:30 AM  
- OUT valid: 03:15 → 03:20 PM  
- OUT earlier than 03:15 → **Early**
""")

attendance_file = st.file_uploader("📁 Upload Attendance File", type=["xlsx"])
bio1 = st.file_uploader("📁 Upload Biometric File – Day 1", type=["xlsx"])
bio2 = st.file_uploader("📁 Upload Biometric File – Day 2", type=["xlsx"])

if st.button("🔍 Compare Files"):
    if not attendance_file or not bio1 or not bio2:
        st.error("⚠️ Please upload all three files before comparing.")
    else:
        with st.spinner("Comparing files... please wait..."):
            try:
                output, df = compare_files(attendance_file.read(), bio1.read(), bio2.read())
            except Exception as e:
                st.error(f"❌ Error while processing files:\n\n{e}")
            else:
                st.success("✅ Comparison completed successfully!")
                st.download_button(
                    label="⬇️ Download Updated Excel File",
                    data=output,
                    file_name="Attendance_with_Status.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.markdown(f"### 🧾 Output Preview — {len(df)} rows")
                st.dataframe(df.head(20))
